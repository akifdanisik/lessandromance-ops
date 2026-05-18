"""Build Less and Romance wholesale order form from Shopify product export."""
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName

SOURCE = "shopify_products_snapshot.json"
OUT = "LessAndRomance_Wholesale_OrderForm.xlsx"

# Discount tiers
FULL_DISCOUNT = 0.50            # buyer hits MOQ → 50% off retail
REDUCED_DISCOUNT = 0.225        # buyer below MOQ → 22.5% off (midpoint of 20-25%)
MOQ_EUR = 10000                 # minimum wholesale order value to qualify for full discount

# Out-of-stock variants we keep on the order form anyway (backorder / hero pieces)
KEEP_OOS_SKUS = {
    "47420466856036",   # Soft Lounge Kimono Grey Melange — STANDART
    "47280364847204",   # Soft Lounge Pant Butter Yellow — XS-S
    "47420530163812",   # Soft Lounge Long Tee Grey Melange — M
}

COLOR_KEYWORDS = [
    "BUTTER YELLOW", "BABY BLUE", "GREY MELANGE", "BITTER COFFEE",
    "DARK MINK", "LIGHT MINK",
    "BURGUNDY", "BLACK", "WHITE", "RED", "MINK", "STONE", "ECRU", "BEIGE",
    "NAVY", "BROWN", "CREAM", "OLIVE", "KHAKI", "PINK", "GREEN", "BLUE",
    "GREY", "GRAY", "YELLOW",
]
SIZE_ORDER = ["XS", "S", "M", "L", "XL", "XS-S", "M-L", "STANDART", "ONE SIZE"]


def parse_color(title: str, color_option_value: str | None) -> tuple[str, str]:
    up = title.upper()
    for c in COLOR_KEYWORDS:
        if up.endswith(" " + c):
            style = title[: -(len(c) + 1)].strip()
            return style, c.title()
    if color_option_value:
        return title, color_option_value.replace("-", " ").title()
    return title, ""


def size_sort_key(s: str) -> tuple[int, str]:
    s_up = s.upper()
    if s_up in SIZE_ORDER:
        return (SIZE_ORDER.index(s_up), s_up)
    return (999, s_up)


def main():
    with open(SOURCE) as f:
        data = json.load(f)
    products = [e["node"] for e in data["products"]["edges"]]

    # Pre-count variants that will land on the Order Form so Order Info totals
    # can reference an exact bounded range (avoids double-counting the TOTALS row).
    variant_count = 0
    for p in products:
        for ve in p["variants"]["edges"]:
            v = ve["node"]
            stock = v.get("inventoryQuantity") or 0
            sku = v.get("sku") or ""
            if stock <= 0 and sku not in KEEP_OOS_SKUS:
                continue
            variant_count += 1
    last_data_row = variant_count + 1  # data starts at row 2

    wb = Workbook()

    brand_font = Font(name="Helvetica", size=22, bold=True, color="1A1A1A")
    label_font = Font(name="Helvetica", size=10, bold=True, color="555555")
    value_fill = PatternFill("solid", fgColor="F5F1EB")
    input_fill = PatternFill("solid", fgColor="FFF9E6")
    thin = Side(style="thin", color="D4CFC4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(name="Helvetica", size=10, bold=True, color="FFFFFF")

    # ============ Sheet 1: Order Info + Pricing tier ============
    ws_cover = wb.active
    ws_cover.title = "Order Info"

    ws_cover["B2"] = "LESS AND ROMANCE"
    ws_cover["B2"].font = brand_font
    ws_cover["B3"] = "Wholesale Order Form — Spring/Summer 2026"
    ws_cover["B3"].font = Font(name="Helvetica", size=11, italic=True, color="555555")
    ws_cover["B4"] = "Currency: EUR  •  Prices net of VAT"
    ws_cover["B4"].font = Font(name="Helvetica", size=9, color="888888")

    cover_fields = [
        "Buyer / Boutique", "Contact name", "Email", "Phone",
        "Shipping address", "VAT / Tax ID", "Preferred delivery date",
        "Payment terms", "Notes",
    ]
    r = 6
    for label in cover_fields:
        ws_cover.cell(row=r, column=2, value=label).font = label_font
        c = ws_cover.cell(row=r, column=3, value="")
        c.fill = value_fill
        c.border = border
        ws_cover.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws_cover.row_dimensions[r].height = 22
        r += 1

    # ----- Pricing Tier block -----
    r += 1
    ws_cover.cell(row=r, column=2, value="PRICING TIER").font = Font(name="Helvetica", size=12, bold=True)
    r += 1

    # Static parameters (editable)
    moq_row = r
    ws_cover.cell(row=r, column=2, value="Minimum Order Quantity (MOQ)").font = label_font
    c = ws_cover.cell(row=r, column=3, value=MOQ_EUR)
    c.fill = input_fill
    c.border = border
    c.number_format = '#,##0 €'
    c.font = Font(bold=True)
    ws_cover.cell(row=r, column=4, value="Wholesale subtotal threshold for full discount").font = Font(italic=True, size=9, color="888888")
    r += 1

    full_disc_row = r
    ws_cover.cell(row=r, column=2, value="Full discount (MOQ met)").font = label_font
    c = ws_cover.cell(row=r, column=3, value=FULL_DISCOUNT)
    c.fill = input_fill
    c.border = border
    c.number_format = '0.0%'
    c.font = Font(bold=True)
    ws_cover.cell(row=r, column=4, value="Applied to retail when order ≥ MOQ").font = Font(italic=True, size=9, color="888888")
    r += 1

    red_disc_row = r
    ws_cover.cell(row=r, column=2, value="Reduced discount (below MOQ)").font = label_font
    c = ws_cover.cell(row=r, column=3, value=REDUCED_DISCOUNT)
    c.fill = input_fill
    c.border = border
    c.number_format = '0.0%'
    c.font = Font(bold=True)
    ws_cover.cell(row=r, column=4, value="Applied when order < MOQ (range 20–25%)").font = Font(italic=True, size=9, color="888888")
    r += 1

    # Qualifying subtotal = SUMPRODUCT(Qty × Retail × FullDiscount) — the value of the order AT THE FULL WHOLESALE PRICE.
    # That's what we check against MOQ to decide which tier applies.
    r += 1
    qual_row = r
    ws_cover.cell(row=r, column=2, value="Qualifying subtotal (@ full discount)").font = label_font
    c = ws_cover.cell(row=r, column=3,
        value=f"=SUMPRODUCT('Order Form'!I2:I{last_data_row},'Order Form'!F2:F{last_data_row})*$C${full_disc_row}")
    c.fill = value_fill
    c.border = border
    c.number_format = '#,##0.00 €'
    r += 1

    applied_row = r
    ws_cover.cell(row=r, column=2, value="→ Applied discount rate").font = Font(bold=True)
    c = ws_cover.cell(row=r, column=3,
        value=f"=IF(C{qual_row}>=C{moq_row},C{full_disc_row},C{red_disc_row})")
    c.fill = PatternFill("solid", fgColor="E8E2D5")
    c.border = border
    c.number_format = '0.0%'
    c.font = Font(bold=True, size=11)
    r += 1

    status_row = r
    ws_cover.cell(row=r, column=2, value="→ MOQ status").font = Font(bold=True)
    c = ws_cover.cell(row=r, column=3,
        value=f'=IF(C{qual_row}>=C{moq_row},"MOQ MET — full discount applied","BELOW MOQ — reduced discount applied")')
    c.fill = PatternFill("solid", fgColor="E8E2D5")
    c.border = border
    c.font = Font(bold=True, size=11)
    r += 1

    # ----- Define named ranges so Order Form sheet can reference them -----
    wb.defined_names["AppliedDiscount"] = DefinedName(
        name="AppliedDiscount",
        attr_text=f"'Order Info'!$C${applied_row}")
    wb.defined_names["MOQ"] = DefinedName(
        name="MOQ", attr_text=f"'Order Info'!$C${moq_row}")
    wb.defined_names["FullDiscount"] = DefinedName(
        name="FullDiscount", attr_text=f"'Order Info'!$C${full_disc_row}")

    # ----- Order Totals block (recomputed live) -----
    r += 1
    ws_cover.cell(row=r, column=2, value="ORDER TOTALS").font = Font(name="Helvetica", size=12, bold=True)
    r += 1
    # Bounded ranges — never reference whole columns, because the TOTALS row
    # at the bottom of the Order Form already contains a sum and would double-count.
    qty_range = f"'Order Form'!I2:I{last_data_row}"
    retail_range = f"'Order Form'!F2:F{last_data_row}"
    line_range = f"'Order Form'!J2:J{last_data_row}"
    totals = [
        ("Total units", f"=SUM({qty_range})", '#,##0'),
        ("Subtotal at retail (EUR)", f"=SUMPRODUCT({qty_range},{retail_range})", '#,##0.00 €'),
        ("Wholesale subtotal (EUR)", f"=SUM({line_range})", '#,##0.00 €'),
        ("Effective discount off retail", f"=C{applied_row}", '0.0%'),
        ("Amount above / below MOQ", f"=SUM({line_range})-C{moq_row}", '#,##0.00 €'),
    ]
    for label, formula, fmt in totals:
        ws_cover.cell(row=r, column=2, value=label).font = label_font
        cell = ws_cover.cell(row=r, column=3, value=formula)
        cell.fill = value_fill
        cell.border = border
        cell.number_format = fmt
        r += 1

    ws_cover.column_dimensions["A"].width = 2
    ws_cover.column_dimensions["B"].width = 32
    ws_cover.column_dimensions["C"].width = 22
    for col in ["D", "E", "F", "G"]:
        ws_cover.column_dimensions[col].width = 18

    # ============ Sheet 2: Order Form ============
    ws = wb.create_sheet("Order Form")
    headers = ["#", "Style", "Color", "Size", "SKU", "Retail (EUR)",
               "Wholesale (EUR)", "Stock", "Qty", "Line total (EUR)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    row = 2
    idx = 1
    for p in products:
        title = p["title"]
        opts_by_name = {o["name"]: o["values"] for o in p["options"]}
        color_opt = opts_by_name.get("Color", [None])[0]
        style, color_from_title = parse_color(title, color_opt)

        variants = []
        for ve in p["variants"]["edges"]:
            v = ve["node"]
            stock = v.get("inventoryQuantity") or 0
            sku = v.get("sku") or ""
            # Drop OOS variants unless on the keep-list
            if stock <= 0 and sku not in KEEP_OOS_SKUS:
                continue
            so = {x["name"]: x["value"] for x in v["selectedOptions"]}
            size = so.get("SIZE") or so.get("Size") or "STANDART"
            color_v = so.get("Color") or color_from_title
            if color_v and "-" in color_v:
                color_v = color_v.replace("-", " ").title()
            variants.append({
                "size": size,
                "color": color_v or color_from_title or "—",
                "sku": sku,
                "price": float(v["price"]),
                "stock": stock,
            })
        if not variants:
            continue
        variants.sort(key=lambda x: (x["color"], size_sort_key(x["size"])))

        for v in variants:
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=style)
            ws.cell(row=row, column=3, value=v["color"])
            ws.cell(row=row, column=4, value=v["size"])
            ws.cell(row=row, column=5, value=v["sku"])
            ws.cell(row=row, column=6, value=v["price"]).number_format = '#,##0.00'
            # Wholesale = retail × (1 - applied discount)  — recomputes live based on MOQ tier
            ws.cell(row=row, column=7, value=f"=F{row}*(1-AppliedDiscount)").number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=v["stock"])
            qty_cell = ws.cell(row=row, column=9, value=0)
            qty_cell.fill = input_fill
            qty_cell.border = border
            ws.cell(row=row, column=10, value=f"=I{row}*G{row}").number_format = '#,##0.00'
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name="Helvetica", size=10)
                cell.alignment = Alignment(vertical="center",
                    horizontal="center" if col in (1, 4, 6, 7, 8, 9, 10) else "left")
                if col != 9:
                    cell.border = border
            if idx % 2 == 0:
                for col in range(1, 11):
                    if col == 9:
                        continue
                    existing = ws.cell(row=row, column=col).fill
                    if existing.fgColor.rgb in (None, "00000000"):
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FAF8F4")
            row += 1
        idx += 1

    last_data_row = row - 1
    total_row = row + 1
    ws.cell(row=total_row, column=8, value="TOTALS").font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{last_data_row})").font = Font(bold=True)
    ws.cell(row=total_row, column=9).number_format = '#,##0'
    ws.cell(row=total_row, column=10, value=f"=SUM(J2:J{last_data_row})").font = Font(bold=True)
    ws.cell(row=total_row, column=10).number_format = '#,##0.00 €'
    for col in (8, 9, 10):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="E8E2D5")
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).alignment = Alignment(horizontal="center")

    # MOQ note row
    ws.cell(row=total_row + 1, column=2,
        value="MOQ €10,000 wholesale subtotal. Below MOQ, discount drops to 20–25% off retail (see Order Info).")
    ws.cell(row=total_row + 1, column=2).font = Font(italic=True, size=9, color="888888")

    widths = [5, 36, 18, 10, 16, 13, 16, 9, 8, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Conditional formatting on stock
    red_fill = PatternFill("solid", fgColor="F4D6D6")
    ws.conditional_formatting.add(f"H2:H{last_data_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=red_fill))

    # Highlight rows where qty > stock
    overstock_fill = PatternFill("solid", fgColor="FCE4A6")
    ws.conditional_formatting.add(f"I2:I{last_data_row}",
        FormulaRule(formula=[f"AND(I2>0,I2>H2)"], fill=overstock_fill))

    ws.auto_filter.ref = f"A1:J{last_data_row}"

    # ============ Sheet 3: Style Summary ============
    ws_sum = wb.create_sheet("Style Summary")
    sum_headers = ["Style", "Color", "Variants", "Sizes available",
                   "Retail EUR (range)", "Total stock"]
    for i, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center")

    groups = {}
    for p in products:
        title = p["title"]
        opts_by_name = {o["name"]: o["values"] for o in p["options"]}
        color_opt = opts_by_name.get("Color", [None])[0]
        style, color_from_title = parse_color(title, color_opt)
        for ve in p["variants"]["edges"]:
            v = ve["node"]
            stock = v.get("inventoryQuantity") or 0
            sku = v.get("sku") or ""
            if stock <= 0 and sku not in KEEP_OOS_SKUS:
                continue
            so = {x["name"]: x["value"] for x in v["selectedOptions"]}
            size = so.get("SIZE") or so.get("Size") or "STANDART"
            color_v = so.get("Color") or color_from_title or "—"
            if color_v and "-" in color_v:
                color_v = color_v.replace("-", " ").title()
            key = (style, color_v)
            g = groups.setdefault(key, {"sizes": [], "prices": [], "stock": 0})
            g["sizes"].append(size)
            g["prices"].append(float(v["price"]))
            g["stock"] += stock

    r = 2
    for (style, color), g in sorted(groups.items()):
        sizes_sorted = sorted(set(g["sizes"]), key=size_sort_key)
        pmin, pmax = min(g["prices"]), max(g["prices"])
        price_range = f"{pmin:.2f}" if pmin == pmax else f"{pmin:.2f} – {pmax:.2f}"
        ws_sum.cell(row=r, column=1, value=style)
        ws_sum.cell(row=r, column=2, value=color)
        ws_sum.cell(row=r, column=3, value=len(g["sizes"]))
        ws_sum.cell(row=r, column=4, value=", ".join(sizes_sorted))
        ws_sum.cell(row=r, column=5, value=price_range)
        ws_sum.cell(row=r, column=6, value=g["stock"])
        for col in range(1, 7):
            cell = ws_sum.cell(row=r, column=col)
            cell.font = Font(name="Helvetica", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                horizontal="center" if col in (3, 5, 6) else "left")
        r += 1
    sum_widths = [36, 22, 10, 28, 22, 12]
    for i, w in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.row_dimensions[1].height = 26
    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:F{r-1}"

    # ============ Sheet 4: Terms ============
    ws_t = wb.create_sheet("Terms")
    terms = [
        ("LESS AND ROMANCE — Wholesale Terms", 18, True),
        ("", 10, False),
        ("Minimum Order Quantity (MOQ)", 12, True),
        ("• MOQ: €10,000 wholesale subtotal per order", 10, False),
        ("• Orders ≥ €10,000 receive the full 50% off retail discount", 10, False),
        ("• Orders below MOQ are accepted but with a reduced 20–25% off retail discount", 10, False),
        ("• MOQ tier is calculated automatically on the 'Order Info' tab", 10, False),
        ("", 10, False),
        ("Pricing", 12, True),
        ("• All prices in EUR, net of VAT", 10, False),
        ("• Full wholesale price = 50% off retail (when MOQ is met)", 10, False),
        ("• Reduced wholesale price = 20–25% off retail (when MOQ is not met)", 10, False),
        ("• Retail prices shown are the suggested retail price (SRP)", 10, False),
        ("", 10, False),
        ("Payment", 12, True),
        ("• 50% deposit on order confirmation", 10, False),
        ("• Balance due before shipment", 10, False),
        ("• Bank transfer (IBAN provided on proforma invoice)", 10, False),
        ("", 10, False),
        ("Delivery", 12, True),
        ("• Lead time: 2–4 weeks from confirmed order", 10, False),
        ("• Shipping ex-works Istanbul, or DAP on request", 10, False),
        ("", 10, False),
        ("Returns", 12, True),
        ("• Wholesale orders are final sale", 10, False),
        ("• Defects must be reported within 7 days of receipt", 10, False),
        ("", 10, False),
        ("Contact", 12, True),
        ("• Email: hello@lessandromance.com", 10, False),
        ("• Instagram: @lessandromance", 10, False),
    ]
    for i, (text, size, bold) in enumerate(terms, 1):
        c = ws_t.cell(row=i, column=2, value=text)
        c.font = Font(name="Helvetica", size=size, bold=bold,
                      color="1A1A1A" if bold else "333333")
    ws_t.column_dimensions["A"].width = 2
    ws_t.column_dimensions["B"].width = 90

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Products: {len(products)} | Variants: {last_data_row - 1} | Style-color groups: {len(groups)}")
    print(f"MOQ: €{MOQ_EUR:,} | Full discount: {FULL_DISCOUNT:.0%} | Reduced: {REDUCED_DISCOUNT:.1%}")


if __name__ == "__main__":
    main()
